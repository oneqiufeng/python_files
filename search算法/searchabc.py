import openpyxl
import requests
from bs4 import BeautifulSoup

# 读取 Excel 文件
workbook = openpyxl.load_workbook('your_excel_file.xlsx')
sheet = workbook.active

# 选择要搜索的列，假设为第一列（A列），并且搜索结果将写入第二列（B列）
search_column = 'A'
result_column = 'B'

# 设置搜索引擎和搜索查询
search_engine = 'https://www.bing.com/search?q='

# 遍历每个单元格，执行搜索并提取链接
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
    query = row[0]
    search_url = search_engine + query
    
    # 发送 HTTP 请求并解析 HTML 内容
    response = requests.get(search_url)
    soup = BeautifulSoup(response.text, 'html.parser')
    
    # 提取搜索结果中的第一个链接
    first_link = soup.find('a')['href'] if soup.find('a') else 'No search results found'
    
    # 将链接写入结果列
    sheet[f'{result_column}{sheet.max_row + 1}'] = first_link

# 保存 Excel 文件
workbook.save('your_excel_file_with_links.xlsx')